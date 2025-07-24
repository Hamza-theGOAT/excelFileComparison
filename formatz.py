import openpyxl
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table
from datetime import datetime


def maxRange(ws, r0, r1, c0, c1):
    r0 = r0 if r0 is not None else ws.min_row
    r1 = r1 if r1 is not None else ws.max_row
    c0 = c0 if c0 is not None else ws.min_col
    c1 = c1 if c1 is not None else ws.max_col
    return r0, r1, c0, c1


def Alt_HOI(worksheet):
    column_widths = {}
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value:
                # Calculate the maximum length of each column
                column_widths[cell.column_letter] = max(
                    (column_widths.get(cell.column_letter, 0)),
                    len(str(cell.value))
                )
    # Add a little extra width for aesthetics
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width + 3


def Alt_A(worksheet):
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    return f"A1:{openpyxl.utils.get_column_letter(max_col)}{max_row}"


def Alt_HK(ws, r0, r1, c0, c1):
    maxRange(ws, r0, r1, c0, c1)

    for col in ws.iter_cols(min_row=r0, max_row=r1, min_col=c0, max_col=c1):
        for cell in col:
            cell.number_format = '#,##0.00;(#,##0.00);0'


def Alt_HNN(ws, r0, r1, c0, c1):
    maxRange(ws, r0, r1, c0, c1)

    for col in ws.iter_cols(min_row=r0, max_row=r1, min_col=c0, max_col=c1):
        for cell in col:
            cell.number_format = '###0;(###0);0'


def Alt_HNS(ws, r0, r1, c0, c1):
    maxRange(ws, r0, r1, c0, c1)

    for col in ws.iter_cols(min_row=r0, max_row=r1, min_col=c0, max_col=c1):
        for cell in col:
            # Consider removing this, instead of changing Date to str in working.py
            if isinstance(cell.value, str):
                cell.value = datetime.strptime(cell.value, '%m-%d-%Y')
            cell.number_format = 'MM/DD/YYYY'


def tableFormat(wb):
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.freeze_panes = 'A2'
        for col in ws.iter_cols(max_row=1, max_col=ws.max_column):
            for cell in col:
                cell.alignment = Alignment(horizontal='left')
        ws.sheet_view.showGridLines = False
        ws_data = Alt_A(ws)
        ws.add_table(Table(displayName=f'{sheet}', ref=ws_data))
        ws = Alt_HOI(ws)
